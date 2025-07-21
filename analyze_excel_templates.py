#!/usr/bin/env python3
"""
Analyze LGU Excel Templates
Extract structure and formatting from official RPMES templates
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, PatternFill
import json
import os

def analyze_worksheet(ws, sheet_name):
    """Analyze a worksheet and extract its structure"""
    print(f"\n=== Analyzing {sheet_name} ===")
    
    # Get basic info
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}")
    
    # Analyze cell structure
    structure = {
        'sheet_name': sheet_name,
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'merged_cells': [],
        'headers': [],
        'data_cells': [],
        'formats': {}
    }
    
    # Get merged cells
    for merged_range in ws.merged_cells.ranges:
        structure['merged_cells'].append(str(merged_range))
    
    print(f"Merged Cells: {len(structure['merged_cells'])}")
    
    # Analyze first 20 rows for headers and structure
    for row in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(11, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell_info = {
                    'row': row,
                    'col': col,
                    'value': str(cell.value),
                    'coordinate': cell.coordinate,
                    'font_bold': cell.font.bold if cell.font else False,
                    'font_size': cell.font.size if cell.font else None,
                    'alignment': {
                        'horizontal': cell.alignment.horizontal if cell.alignment else None,
                        'vertical': cell.alignment.vertical if cell.alignment else None
                    },
                    'border': bool(cell.border) if cell.border else False,
                    'fill': bool(cell.fill) if cell.fill else False
                }
                row_data.append(cell_info)
                
                # Check if it's a header
                if cell.font and cell.font.bold:
                    structure['headers'].append(cell_info)
        
        if row_data:
            structure['data_cells'].append(row_data)
            print(f"Row {row}: {len(row_data)} cells with data")
    
    return structure

def analyze_excel_file(filename):
    """Analyze an Excel file and extract all worksheet structures"""
    print(f"\n{'='*60}")
    print(f"ANALYZING: {filename}")
    print(f"{'='*60}")
    
    try:
        wb = openpyxl.load_workbook(filename)
        print(f"Workbook loaded successfully")
        print(f"Sheets: {wb.sheetnames}")
        
        structures = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            structure = analyze_worksheet(ws, sheet_name)
            structures[sheet_name] = structure
        
        # Save analysis to JSON
        output_file = f"{os.path.splitext(filename)[0]}_analysis.json"
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(structures, f, indent=2, ensure_ascii=False)
        
        print(f"\nAnalysis saved to: {output_file}")
        return structures
        
    except Exception as e:
        print(f"Error analyzing {filename}: {e}")
        return None

def main():
    """Main analysis function"""
    print("LGU Excel Template Analysis")
    print("=" * 50)
    
    # Analyze Input Forms
    input_file = "MSWDO-2025-RPMES-Input-Forms-1-4 (1).xlsx"
    if os.path.exists(input_file):
        input_structures = analyze_excel_file(input_file)
    else:
        print(f"Input file not found: {input_file}")
        input_structures = None
    
    # Analyze Output Forms
    output_file = "RPMES-Output-Forms-5-11.xlsx"
    if os.path.exists(output_file):
        output_structures = analyze_excel_file(output_file)
    else:
        print(f"Output file not found: {output_file}")
        output_structures = None
    
    # Generate summary
    print(f"\n{'='*60}")
    print("ANALYSIS SUMMARY")
    print(f"{'='*60}")
    
    if input_structures:
        print(f"Input Forms Analysis:")
        for sheet_name, structure in input_structures.items():
            print(f"  - {sheet_name}: {structure['max_row']} rows, {structure['max_col']} cols, {len(structure['merged_cells'])} merged cells")
    
    if output_structures:
        print(f"Output Forms Analysis:")
        for sheet_name, structure in output_structures.items():
            print(f"  - {sheet_name}: {structure['max_row']} rows, {structure['max_col']} cols, {len(structure['merged_cells'])} merged cells")
    
    print(f"\nAnalysis complete! Check the generated JSON files for detailed structure.")

if __name__ == "__main__":
    main() 